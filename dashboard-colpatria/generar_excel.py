"""
Generar Excel con Lista COMPLETA de Entidades y Ciudades Normalizadas
====================================================================
Este script genera un Excel con TODAS las entidades y ciudades normalizadas
junto con SUS VARIANTES COMPLETAS.

Ejecutar: python generar_excel.py
Resultado: reporte_normalizacion_completo.xlsx
"""

import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Rutas
DATA_DIR = os.path.join(os.path.dirname(__file__), "dist", "data")
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), "reporte_normalizacion_completo.xlsx")

def cargar_datos():
    """Carga los datos de variantes"""
    print("📂 Cargando datos...")
    with open(os.path.join(DATA_DIR, "variantes_entidad.json"), "r", encoding="utf-8") as f:
        entidades = json.load(f)
    
    with open(os.path.join(DATA_DIR, "variantes_ciudad.json"), "r", encoding="utf-8") as f:
        ciudades = json.load(f)
    
    print(f"   - Entidades: {len(entidades):,}")
    print(f"   - Ciudades: {len(ciudades):,}")
    
    return entidades, ciudades

def crear_dataframe_ciudades(ciudades):
    """Crea un DataFrame con todas las ciudades y sus variantes"""
    print("🔄 Procesando ciudades...")
    
    rows = []
    # Ordenar ciudades por cantidad total (mayor a menor)
    ciudades_ordenadas = sorted(
        ciudades.items(), 
        key=lambda x: sum(v.get('cantidad', 0) for v in x[1]), 
        reverse=True
    )
    
    for normalized_name, variants in ciudades_ordenadas:
        # Ordenar variantes por cantidad
        variants_sorted = sorted(variants, key=lambda x: x.get('cantidad', 0), reverse=True)
        
        for variant in variants_sorted:
            rows.append({
                'Ciudad Normalizada': normalized_name,
                'Variante Original': variant.get('original', ''),
                'Cantidad': variant.get('cantidad', 0),
                'Entidades Diferentes': variant.get('entidades', 0)
            })
    
    df = pd.DataFrame(rows)
    print(f"   - Total filas: {len(df):,}")
    return df

def crear_dataframe_entidades(entidades):
    """Crea un DataFrame con todas las entidades y sus variantes"""
    print("🔄 Procesando entidades...")
    
    rows = []
    # Ordenar entidades por cantidad total (mayor a menor)
    entidades_ordenadas = sorted(
        entidades.items(), 
        key=lambda x: sum(v.get('cantidad', 0) for v in x[1]), 
        reverse=True
    )
    
    for normalized_name, variants in entidades_ordenadas:
        # Ordenar variantes por cantidad
        variants_sorted = sorted(variants, key=lambda x: x.get('cantidad', 0), reverse=True)
        
        for variant in variants_sorted:
            rows.append({
                'Entidad Normalizada': normalized_name,
                'Variante Original': variant.get('original', ''),
                'Cantidad': variant.get('cantidad', 0),
                'Ciudades Diferentes': variant.get('ciudades', '')
            })
    
    df = pd.DataFrame(rows)
    print(f"   - Total filas: {len(df):,}")
    return df

def crear_resumen(entidades, ciudades):
    """Crea DataFrame de resumen"""
    print("🔄 Creando resumen...")
    
    # Calcular estadísticas
    total_entidades = len(entidades)
    total_ciudades = len(ciudades)
    total_var_entidades = sum(len(v) for v in entidades.values())
    total_var_ciudades = sum(len(v) for v in ciudades.values())
    
    # Cantidad total de registros
    total_reg_entidades = sum(
        sum(v.get('cantidad', 0) for v in variants) 
        for variants in entidades.values()
    )
    total_reg_ciudades = sum(
        sum(v.get('cantidad', 0) for v in variants) 
        for variants in ciudades.values()
    )
    
    rows = [
        ['RESUMEN DE NORMALIZACIÓN', '', '', '', ''],
        ['', '', '', '', ''],
        ['Elemento', 'Normalizados', 'Total Variantes', 'Registros Totales', '% Variantes'],
        ['Entidades', total_entidades, total_var_entidades, total_reg_entidades, 
         f"{total_var_entidades/total_entidades:.1f}%"],
        ['Ciudades', total_ciudades, total_var_ciudades, total_reg_ciudades,
         f"{total_var_ciudades/total_ciudades:.1f}%"],
        ['TOTAL', total_entidades + total_ciudades, 
         total_var_entidades + total_var_ciudades,
         total_reg_entidades + total_reg_ciudades, ''],
    ]
    
    return pd.DataFrame(rows)

def estilos_excel(ws, es_resumen=False):
    """Aplica estilos a la hoja de Excel"""
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1d23", end_color="1a1d23", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1a1d23")
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Ajustar ancho de columnas
    if es_resumen:
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
    else:
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
    
    # Aplicar estilos a заголовки
    for cell in ws[1]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    # Alineación para el resto
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = normal_font
            cell.border = border
            if cell.column == 1 or cell.column == 2:  # Nombres
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:  # Números
                cell.alignment = Alignment(horizontal='right', vertical='center')

def generar_excel(df_entidades, df_ciudades, df_resumen):
    """Genera el archivo Excel completo"""
    print("📊 Generando Excel...")
    
    wb = Workbook()
    
    # Hoja 1: Resumen
    print("   - Agregando hoja Resumen...")
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    for row_idx, row_data in enumerate(df_resumen.values, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
    
    # Estilos especiales para resumen
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 18
    ws_resumen.column_dimensions['D'].width = 18
    ws_resumen.column_dimensions['E'].width = 15
    
    # Título en negrita
    for cell in ws_resumen[1]:
        if cell.value:
            cell.font = Font(bold=True, size=14, color="1a1d23")
    
    # Encabezado de tabla
    for cell in ws_resumen[4]:
        if cell.value:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="1a1d23", end_color="1a1d23", fill_type="solid")
    
    # Hoja 2: Todas las Ciudades
    print("   - Agregando hoja Ciudades...")
    ws_ciudades = wb.create_sheet("Ciudades - Todas")
    
    for row_idx, row_data in enumerate(df_ciudades.values, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_ciudades.cell(row=row_idx, column=col_idx, value=value)
    
    estilos_excel(ws_ciudades)
    
    # Hoja 3: Todas las Entidades
    print("   - Agregando hoja Entidades...")
    ws_entidades = wb.create_sheet("Entidades - Todas")
    
    for row_idx, row_data in enumerate(df_entidades.values, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_entidades.cell(row=row_idx, column=col_idx, value=value)
    
    estilos_excel(ws_entidades)
    
    # Hoja 4: Ciudades por cantidad (top)
    print("   - Agregando hoja Top Ciudades...")
    ws_top_ciudades = wb.create_sheet("Top Ciudades")
    
    # Agregar título
    ws_top_ciudades.cell(row=1, column=1, value="TOP 100 CIUDADES POR CANTIDAD DE REGISTROS")
    ws_top_ciudades.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    # Crear lista de ciudades con totales
    ciudades_total = []
    for normalized_name, variants in ciudades.items():
        total = sum(v.get('cantidad', 0) for v in variants)
        num_variantes = len(variants)
        ciudades_total.append({
            'Ciudad Normalizada': normalized_name,
            'Total Registros': total,
            'Número Variantes': num_variantes
        })
    
    # Ordenar por total
    ciudades_total = sorted(ciudades_total, key=lambda x: x['Total Registros'], reverse=True)[:100]
    
    # Agregar datos
    headers = ['Ciudad Normalizada', 'Total Registros', 'Número Variantes']
    for col, header in enumerate(headers, 1):
        cell = ws_top_ciudades.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1a1d23", end_color="1a1d23", fill_type="solid")
    
    for row_idx, city in enumerate(ciudades_total, 4):
        ws_top_ciudades.cell(row=row_idx, column=1, value=city['Ciudad Normalizada'])
        ws_top_ciudades.cell(row=row_idx, column=2, value=city['Total Registros'])
        ws_top_ciudades.cell(row=row_idx, column=3, value=city['Número Variantes'])
    
    ws_top_ciudades.column_dimensions['A'].width = 35
    ws_top_ciudades.column_dimensions['B'].width = 18
    ws_top_ciudades.column_dimensions['C'].width = 18
    
    # Hoja 5: Entidades por cantidad (top)
    print("   - Agregando hoja Top Entidades...")
    ws_top_entidades = wb.create_sheet("Top Entidades")
    
    # Agregar título
    ws_top_entidades.cell(row=1, column=1, value="TOP 100 ENTIDADES POR CANTIDAD DE REGISTROS")
    ws_top_entidades.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    # Crear lista de entidades con totales
    entidades_total = []
    for normalized_name, variants in entidades.items():
        total = sum(v.get('cantidad', 0) for v in variants)
        num_variantes = len(variants)
        entidades_total.append({
            'Entidad Normalizada': normalized_name,
            'Total Registros': total,
            'Número Variantes': num_variantes
        })
    
    # Ordenar por total
    entidades_total = sorted(entidades_total, key=lambda x: x['Total Registros'], reverse=True)[:100]
    
    # Agregar datos
    for col, header in enumerate(headers, 1):
        cell = ws_top_entidades.cell(row=3, column=col, value=header.replace('Ciudad', 'Entidad'))
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1a1d23", end_color="1a1d23", fill_type="solid")
    
    for row_idx, ent in enumerate(entidades_total, 4):
        ws_top_entidades.cell(row=row_idx, column=1, value=ent['Entidad Normalizada'])
        ws_top_entidades.cell(row=row_idx, column=2, value=ent['Total Registros'])
        ws_top_entidades.cell(row=row_idx, column=3, value=ent['Número Variantes'])
    
    ws_top_entidades.column_dimensions['A'].width = 45
    ws_top_entidades.column_dimensions['B'].width = 18
    ws_top_entidades.column_dimensions['C'].width = 18
    
    # Guardar
    wb.save(OUTPUT_XLSX)
    
    size = os.path.getsize(OUTPUT_XLSX) / 1024
    print(f"✅ Excel generado: {OUTPUT_XLSX} ({size:.1f} KB)")

if __name__ == "__main__":
    print("=" * 60)
    print("GENERADOR DE REPORTE EXCEL - NORMALIZACIÓN COMPLETA")
    print("=" * 60)
    
    # Cargar datos
    entidades, ciudades = cargar_datos()
    
    # Crear DataFrames
    df_resumen = crear_resumen(entidades, ciudades)
    df_ciudades = crear_dataframe_ciudades(ciudades)
    df_entidades = crear_dataframe_entidades(entidades)
    
    # Generar Excel
    generar_excel(df_entidades, df_ciudades, df_resumen)
    
    print("\n🎉 ¡Proceso completado!")
    print(f"   Archivo: {OUTPUT_XLSX}")
    print("\n📋 Hojas del Excel:")
    print("   1. Resumen - Estadísticas generales")
    print("   2. Ciudades - Todas las ciudades con TODAS sus variantes")
    print("   3. Entidades - Todas las entidades con TODAS sus variantes")
    print("   4. Top Ciudades - Top 100 ciudades por cantidad")
    print("   5. Top Entidades - Top 100 entidades por cantidad")

